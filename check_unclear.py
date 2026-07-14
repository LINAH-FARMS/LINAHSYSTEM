import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook('C:/Users/Salem Magdy/Downloads/الإدارة الفنية.xlsx')
ws = wb.active
sector_room_map = {
    'الجزورين غرفة 1', 'الجزورين غ 1', 'الجزورين غ 2', 'الجزورين غ 3',
    'الجزورين غ 4', 'الجزورين غ 5', 'الجزورين غ 6',
    'السكن الجديد F3', 'السكن الجديد F6',
    'النخالين غ 1', 'النخالين غ 2',
    'ق 17', 'قطاع 1', 'ق 1', 'ق 3', 'قطاع 3', 'ق 6',
    'ق 21', 'ق 24', 'ق 25', 'ق 26', 'ق 27', 'ق 28', 'ق 29', 'ق 30', 'ق 32', 'ق 33',
}
for r in range(2, ws.max_row + 1):
    name = str(ws.cell(r, 1).value or '').strip()
    code = str(ws.cell(r, 2).value or '').strip()
    loc = str(ws.cell(r, 3).value or '').strip()
    if loc and loc not in sector_room_map:
        print(f"Row {r}: UNMAPPED | name='{name}' | code='{code}' | loc='{loc}'")
