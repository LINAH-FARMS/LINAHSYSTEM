import openpyxl, json, sys, urllib.request
sys.stdout.reconfigure(encoding='utf-8')

apiKey = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImN3cWdoaXF5a29oZWZhZ2dlZGpsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEwMjUyMjEsImV4cCI6MjA5NjYwMTIyMX0.3a3hRcNdmYQCtjYjBroAT6df1T_7oz-XWUeD3wagYw8'

req = urllib.request.Request('https://cwqghiqykohefaggedjl.supabase.co/rest/v1/sync_data?id=eq.alldata&select=data',
    headers={'apikey': apiKey, 'Authorization': f'Bearer {apiKey}'})
with urllib.request.urlopen(req, timeout=15) as r:
    result = json.loads(r.read())
data = result[0]['data']
if isinstance(data, str): data = json.loads(data)
emps = data['employees']

# Read file
wb = openpyxl.load_workbook(r'C:\Users\Salem Magdy\Downloads\الإدارة الفنية (1).xlsx')
ws = wb.active
file_emps = []
skip_locs = {'الإستراحة', 'الإستراحة المهندسين', 'بيفوت 17'}
for r in range(2, ws.max_row + 1):
    name = str(ws.cell(r, 1).value or '').strip()
    code = str(ws.cell(r, 2).value or '').strip()
    loc = str(ws.cell(r, 4).value or '').strip()
    status_raw = str(ws.cell(r, 5).value or '').strip()
    status = 'V' if 'جاز' in status_raw else 'P'
    if name and code and loc not in skip_locs:
        file_emps.append({'name': name, 'code': code, 'loc': loc, 'status': status})

loc_map = {
    'الجزورين غ 1': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 1'),
    'الجزورين غ 2': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 2'),
    'الجزورين غ 3': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 3'),
    'الجزورين غ 4': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 4'),
    'الجزورين غ 5': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 5'),
    'الجزورين غ 6': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 6'),
    'السكن الجديد F3': ('سكن العاملين الجديد 2025 (F)', 'غرفه F3'),
    'السكن الجديد F6': ('سكن العاملين الجديد 2025 (F)', 'غرفه F6'),
    'النخالين غ 1': ('سكن العاملين (سكن النخالين)', 'غرفه نخاليين 1'),
    'النخالين غ 2': ('سكن العاملين (سكن النخالين)', 'غرفه نخاليين 2'),
    'سكن المهندسين A1': ('سكن المهندسين (السكن الجديد)', 'غرفه A1'),
    'سكن المهندسين A3': ('سكن المهندسين (السكن الجديد)', 'غرفه A3'),
    'ق 1': ('سكن القطاعات', 'قطاع رقم ( 1 )'),
    'ق 3': ('سكن القطاعات', 'قطاع رقم ( 3 )'),
    'ق 6': ('سكن القطاعات', 'قطاع رقم ( 6 )'),
    'ق 21': ('سكن القطاعات', 'قطاع رقم ( 21 )'),
    'ق 24': ('سكن القطاعات', 'قطاع رقم ( 24 )'),
    'ق 25': ('سكن القطاعات', 'قطاع رقم ( 25 )'),
    'ق 26': ('سكن القطاعات', 'قطاع رقم ( 26 )'),
    'ق 27': ('سكن القطاعات', 'قطاع رقم ( 27 )'),
    'ق 28': ('سكن القطاعات', 'قطاع رقم ( 28 )'),
    'ق 29': ('سكن القطاعات', 'قطاع رقم ( 29 )'),
    'ق 30': ('سكن القطاعات', 'قطاع رقم ( 30 )'),
    'ق 32': ('سكن القطاعات', 'قطاع رقم ( 32 )'),
    'ق 33': ('سكن القطاعات', 'قطاع رقم ( 33 )'),
}

correct = []
wrong = []
not_found = []

for item in file_emps:
    exp_s, exp_r = loc_map[item['loc']]
    emp = next((e for e in emps if str(e.get('code','')) == item['code']), None)
    if not emp:
        not_found.append(f"  {item['code']:>6} | {item['name'][:35]:<35} | {item['loc']} — NOT IN PROGRAM")
        continue
    act_s = emp.get('sector','')
    act_r = emp.get('room','')
    act_st = emp.get('status','')
    file_st = item['status']
    if act_s == exp_s and act_r == exp_r:
        # Check status too
        if act_st != file_st:
            wrong.append(f"  {item['code']:>6} | {item['name'][:35]:<35} | room OK | status file={file_st} prog={act_st}")
        else:
            correct.append(item['code'])
    else:
        wrong.append(f"  {item['code']:>6} | {item['name'][:35]:<35} | Expected: {exp_s}/{exp_r}")
        wrong.append(f"         | {'':35} | Actual  : {act_s}/{act_r}")

print(f"Total checked: {len(file_emps)}")
print(f"✅ Correct: {len(correct)}")
print(f"❌ Wrong: {len(wrong)//2 if wrong else 0}")
print(f"⚠️ Not in program: {len(not_found)}")

if wrong:
    print(f"\n❌ WRONG ({len(wrong)//2}):")
    for w in wrong: print(w)
if not_found:
    print(f"\n⚠️ NOT FOUND ({len(not_found)}):")
    for n in not_found: print(n)
