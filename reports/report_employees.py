import requests, json, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

SUPABASE_URL = 'https://cwqghiqykohefaggedjl.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImN3cWdoaXF5a29oZWZhZ2dlZGpsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEwMjUyMjEsImV4cCI6MjA5NjYwMTIyMX0.3a3hRcNdmYQCtjYjBroAT6df1T_7oz-XWUeD3wagYw8'

def fetch_data():
    r = requests.get(f'{SUPABASE_URL}/rest/v1/sync_data?id=eq.alldata&select=data', headers={'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'})
    d = r.json()
    raw = d[0]['data']
    if isinstance(raw, str): raw = json.loads(raw)
    return raw

def make_employee_report(data):
    emps = data.get('employees', [])
    vacs = data.get('vacations', [])
    excl = data.get('excludedEmployees', [])
    total = len(emps)
    p_count = sum(1 for e in emps if e.get('status') == 'P')
    v_count = sum(1 for e in emps if e.get('status') == 'V')
    depts = {}
    for e in emps:
        d = e.get('dept', e.get('department', 'غير محدد'))
        depts[d] = depts.get(d, 0) + 1
    contracts = {}
    for e in emps:
        c = e.get('contract', 'غير محدد')
        contracts[c] = contracts.get(c, 0) + 1
    wb = Workbook()
    ws = wb.active
    ws.title = 'Employees'
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill('solid', fgColor='1B5E20')
    thin = Side(style='thin'); border = Border(top=thin, left=thin, right=thin, bottom=thin)
    headers = ['القسم', 'العدد', 'النسبة']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
        cell.border = border
    headers2 = ['المجموع', total, '100%']
    for c, v in enumerate(headers2, 1):
        cell = ws.cell(row=2, column=c, value=v)
        cell.font = Font(bold=True, size=11); cell.border = border; cell.alignment = Alignment(horizontal='center')
    ws.cell(row=3, column=1, value='P (متواجد)').border = border; ws.cell(row=3, column=2, value=p_count).border = border; ws.cell(row=3, column=3, value=f'{round(p_count/total*100,1)}%').border = border
    ws.cell(row=4, column=1, value='V (إجازات)').border = border; ws.cell(row=4, column=2, value=v_count).border = border; ws.cell(row=4, column=3, value=f'{round(v_count/total*100,1)}%').border = border
    r = 6
    ws.cell(row=r, column=1, value='توزيع الإدارات').font = Font(bold=True, size=12, color='1B5E20'); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    for d, cnt in sorted(depts.items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=d).border = border; ws.cell(row=r, column=2, value=cnt).border = border; ws.cell(row=r, column=3, value=f'{round(cnt/total*100,1)}%').border = border; r += 1
    r += 1
    ws.cell(row=r, column=1, value='توزيع التعاقدات').font = Font(bold=True, size=12, color='1B5E20'); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    for c, cnt in sorted(contracts.items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=c).border = border; ws.cell(row=r, column=2, value=cnt).border = border; ws.cell(row=r, column=3, value=f'{round(cnt/total*100,1)}%').border = border; r += 1
    for c in range(1, 4): ws.column_dimensions[chr(64+c)].width = 20
    ws.column_dimensions['A'].width = 30
    wb.save('C:\\Users\\Salem Magdy\\Desktop\\تقارير_الموظفين.xlsx')
    print(f'[تم حفظ تقرير الموظفين] {total} موظف')

if __name__ == '__main__':
    data = fetch_data()
    make_employee_report(data)
