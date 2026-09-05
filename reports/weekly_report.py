import requests, json, sys, os, re, argparse
from datetime import datetime, timedelta
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPABASE_URL = 'https://cwqghiqykohefaggedjl.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImN3cWdoaXF5a29oZWZhZ2dlZGpsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEwMjUyMjEsImV4cCI6MjA5NjYwMTIyMX0.3a3hRcNdmYQCtjYjBroAT6df1T_7oz-XWUeD3wagYw8'

# السحابة القديمة — إن كانت مفعلة/مسدودة أجرها، نقرأ منها بالمفتاح القديم
# (حاليًا ترجع 402، لكن الصفوف العادية في السحابة الجديدة هي نسخة منها)
OLD_SUPABASE_URL = 'https://idejmgmftmrniviftcce.supabase.co'
OLD_SUPABASE_KEY = 'sb_publishable_AvMTa-zmQ4hgA1hJNpYc3g_gu8rlirz'

_ARABIC_DIGITS = str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')
_EMOJI_PATTERN = re.compile(r'[\U00002600-\U000027BF\U0001F300-\U0001FAFF\U0000FE00-\U0000FE0F]')

def norm_date(s):
    if not s or not isinstance(s, str): return s or ''
    s = s.replace('\u200f', '').replace('\u200e', '').strip()
    s = s.translate(_ARABIC_DIGITS)
    s = s[:10]
    parts = s.split('/')
    if len(parts) == 3:
        d, m, y = parts
        if len(y) == 4:
            return f'{y}-{int(m):02d}-{int(d):02d}'
    return s

def strip_emoji(s):
    return _EMOJI_PATTERN.sub('', s).strip() if isinstance(s, str) else (s or '')

LOCAL_DATA_FILES = [
    'C:\\Users\\Salem Magdy\\Desktop\\New folder\\linah_merged_cloud.json',
    'C:\\Users\\Salem Magdy\\Desktop\\New folder\\linah_cloud_final.json',
    'C:\\Users\\Salem Magdy\\Desktop\\New folder\\linah_working_project_full.json',
]

_LOCAL_DATA = None
def _load_local_data():
    global _LOCAL_DATA
    if _LOCAL_DATA is not None:
        return _LOCAL_DATA
    merged = {}
    for _p in LOCAL_DATA_FILES:
        try:
            with open(_p, encoding='utf-8') as f:
                _d = json.load(f)
        except Exception as e:
            print(f'[تحذير] تعذر قراءة الملف المحلي {_p}: {e}')
            continue
        if not isinstance(_d, dict):
            continue
        for _k, _v in _d.items():
            if not isinstance(_v, list):
                continue
            # دمج السجلات مع إزالة التكرار لكل كيان (بدون فقدان أي سجل من أي ملف)
            _existing = merged.get(_k)
            if _existing is None:
                merged[_k] = list(_v)
            else:
                _seen = set(_json_key(x) for x in _existing)
                for _x in _v:
                    if _json_key(_x) not in _seen:
                        _existing.append(_x)
                        _seen.add(_json_key(_x))
    _LOCAL_DATA = merged
    return merged

def _json_key(x):
    try:
        return json.dumps(x, ensure_ascii=False, sort_keys=True)
    except Exception:
        return str(x)

_LOCAL_ENTITY_MAP = {
    'ent:employees': 'employees',
    'ent:hospitalities': 'hospitalities',
    'ent:bakeryProductions': 'bakeryProductions',
    'ent:bakeryContractorSupplies': 'bakeryContractorSupplies',
    'ent:mealLogs': 'mealLogs',
    'ent:maintenanceRecords': 'maintenanceRecords',
    'ent:septicRecords': 'septicRecords',
    'ent:teaSugarBatches': 'teaSugarBatches',
    'ent:teaSugarDisbursements': 'teaSugarDisbursements',
    'ent:dailyStats': 'dailyStats',
    'ent:mealWaste': 'mealWaste',
    'ent:syncDeletions': 'syncDeletions',
    'incident_reports': 'incident_reports',
    'meal_waste_entries': 'mealWaste',
}

def fetch_row_local(row_id):
    data = _load_local_data()
    key = _LOCAL_ENTITY_MAP.get(row_id, row_id)
    v = data.get(key)
    if v is None:
        # try common alternate names
        for k, val in data.items():
            if key.lower() in k.lower() or k.lower() in key.lower():
                return val if isinstance(val, list) else []
        return []
    return v if isinstance(v, list) else []

def fetch_row(row_id):
    try:
        r = requests.get(f'{SUPABASE_URL}/rest/v1/sync_data', params={'id': 'eq.' + row_id}, headers={'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'}, timeout=20)
        j = r.json()
        if isinstance(j, list) and len(j) > 0:
            raw = j[0].get('data')
            return json.loads(raw) if isinstance(raw, str) else raw
        return []
    except Exception as e:
        print(f'[تحذير] فشل سحب {row_id} من السحابة: {e}')
        return []

def fetch_row_old(row_id):
    """محاولة سحب من السحابة القديمة (قد ترجع 402 إذا كانت مسدودة)."""
    try:
        r = requests.get(f'{OLD_SUPABASE_URL}/rest/v1/sync_data', params={'id': 'eq.' + row_id}, headers={'apikey': OLD_SUPABASE_KEY, 'Authorization': f'Bearer {OLD_SUPABASE_KEY}'}, timeout=15)
        j = r.json()
        if isinstance(j, list) and len(j) > 0:
            raw = j[0].get('data')
            return json.loads(raw) if isinstance(raw, str) else raw
        return []
    except Exception:
        return []

def _entity_merge_key(entity, x):
    if not isinstance(x, dict): return _json_key(x)
    if entity == 'bakeryContractorSupplies': return f"{x.get('name','')}|{x.get('date','')}|{x.get('count','')}|{x.get('price','')}"
    if x.get('id') or x.get('_id'): return str(x.get('id') or x.get('_id'))
    if entity == 'employees': return f"code:{x.get('code','')}|name:{x.get('name','')}"
    if entity == 'mealLogs': return f"date:{norm_date(str(x.get('date','')))}"
    if entity == 'septicRecords': return f"{norm_date(str(x.get('date','')))}|{x.get('name','')}|{x.get('trips','')}|{x.get('quantity','')}"
    if entity == 'maintenanceRecords': return f"{norm_date(str(x.get('date','')))}|{x.get('task','')}|{x.get('tech','')}"
    if entity == 'hospitalities': return f"{x.get('name','')}|{x.get('arrival','')}|{x.get('type','')}"
    if entity in ('dailyStats',): return f"date:{x.get('date','')}"
    if entity == 'mealWaste': return f"{x.get('date','')}|{x.get('meal','')}|{x.get('chef','')}|{x.get('source','')}|{x.get('createdAt','')}"
    if entity == 'bakeryProductions': return f"{x.get('date','')}|{x.get('breadCount','')}|{x.get('flourUsed','')}"
    if entity == 'bakeryContractorSupplies': return f"{x.get('name','')}|{x.get('date','')}|{x.get('count','')}|{x.get('price','')}"
    if entity == 'teaSugarDisbursements': return f"{x.get('id','')}|{x.get('date','')}|{x.get('period','')}|{x.get('empCode','')}"
    if entity == 'teaSugarBatches': return f"{x.get('date','')}|{x.get('name','')}"
    return _json_key(x)

def fetch_row_all(entity, row_id):
    """جمع نفس الكيان من السحابتين + الملفات المحلية مع إزالة التكرار.
    صف ent: (البيانات الحية) + الصف العادي (نسخة السحابة القديمة المنقولة)
    + الملفات المحلية — حتى لا يُفقد أي سجل أثناء انتقال السحابات."""
    out = []
    plain = row_id[4:] if row_id.startswith('ent:') else row_id
    for rid in (row_id, plain):
        v = fetch_row(rid)
        if isinstance(v, list): out += v
    v = fetch_row_old(row_id)
    if isinstance(v, list): out += v
    v = fetch_row_old(plain)
    if isinstance(v, list): out += v
    v = fetch_row_local(plain)
    if isinstance(v, list): out += v
    seen = set(); dedup = []
    for x in out:
        k = _entity_merge_key(entity, x)
        if k in seen: continue
        seen.add(k); dedup.append(x)
    return dedup

def report_window():
    # دائمًا آخر 7 أيام كاملة تنتهي بأمس — مهما كان يوم التشغيل
    # (لو اتشتغل السبت يعطي السبت→الجمعة الماضية، ولو اتشتغل أي يوم تاني
    #  لا يفقد آخر 3 أيام لأن النطاق يتبع آخر يوم مكتمل فعلًا)
    end = datetime.now() - timedelta(days=1)
    return end - timedelta(days=6), end

def fmt(d):
    return d.strftime('%Y-%m-%d')

def style_sheet(ws, headers, rows, col_widths=None, title=None):
    thin = Side(style='thin'); border = Border(top=thin, left=thin, right=thin, bottom=thin)
    alt_fill = PatternFill('solid', fgColor='F5F5F5')
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13, color='1B5E20')
    r = 2 if title else 1
    hf = Font(bold=True, color='FFFFFF', size=10)
    hfill = PatternFill('solid', fgColor='1B5E20')
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = border
    for ri, row_data in enumerate(rows, r + 1):
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=c, value=v)
            cell.border = border; cell.alignment = Alignment(horizontal='center')
            if ri % 2 == 0: cell.fill = alt_fill
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

def norm_filter(arr, key):
    return [x for x in arr if norm_date(x.get(key,'')) >= fmt(start) and norm_date(x.get(key,'')) <= fmt(end)]

print('[جاري سحب البيانات...]')
_ENT_IDS = [
    ('employees', 'ent:employees'),
    ('hospitalities', 'ent:hospitalities'),
    ('bakeryProductions', 'ent:bakeryProductions'),
    ('bakeryContractorSupplies', 'ent:bakeryContractorSupplies'),
    ('mealLogs', 'ent:mealLogs'),
    ('maintenanceRecords', 'ent:maintenanceRecords'),
    ('septicRecords', 'ent:septicRecords'),
    ('teaSugarBatches', 'ent:teaSugarBatches'),
    ('teaSugarDisbursements', 'ent:teaSugarDisbursements'),
    ('dailyStats', 'ent:dailyStats'),
    ('mealWaste', 'ent:mealWaste'),
    ('syncDeletions', 'ent:syncDeletions'),
]
data = {k: fetch_row_all(k, rid) for k, rid in _ENT_IDS}
print(f'[تم السحب] الكيانات (سحابتين + محلي): {", ".join(f"{k}={len(v)}" for k, v in data.items())}')

sync_del = data.get('syncDeletions', [])
def filt(arr, entity):
    if not sync_del or not arr: return arr
    keys = set(d['key'] for d in sync_del if d.get('entity') == entity)
    if not keys: return arr
    key_fns = {
        'employees': lambda e: str(e.get('id','')) or e.get('code') or e.get('name'),
        'hospitalities': lambda h: f"{h.get('name','')}|{h.get('arrival','')}|{h.get('type','')}",
        'maintenanceRecords': lambda m: f"{m.get('category','')}|{m.get('task','')}|{m.get('date','')}",
        'bakeryProductions': lambda p: f"{p.get('date','')}|{p.get('breadCount','')}",
        'bakeryContractorSupplies': lambda s: f"{s.get('name','')}|{s.get('date','')}|{s.get('count','')}",
        'septicRecords': lambda s: f"{s.get('date','')}|{s.get('name',s.get('sector',''))}|{s.get('trips',s.get('quantity',''))}",
        'teaSugarDisbursements': lambda t: f"{t.get('date','')}|{t.get('period',t.get('type',''))}|{t.get('empCode','')}|{t.get('teaPacks',t.get('quantity',''))}|{t.get('sugarKg','')}",
        'incident_reports': lambda r: str(r.get('id','')) or r.get('name','')
    }
    fn = key_fns.get(entity)
    if not fn: return arr
    return [x for x in arr if fn(x) not in keys]

for ent in ['employees','hospitalities','maintenanceRecords','bakeryProductions','bakeryContractorSupplies','septicRecords','teaSugarDisbursements','incident_reports']:
    if data.get(ent): data[ent] = filt(data[ent], ent)

parser = argparse.ArgumentParser()
parser.add_argument('--from', dest='from_date')
parser.add_argument('--to', dest='to_date')
args = parser.parse_args()

if args.from_date and args.to_date:
    start = datetime.strptime(args.from_date, '%Y-%m-%d')
    end = datetime.strptime(args.to_date, '%Y-%m-%d')
else:
    start, end = report_window()
print(f'[الفترة] {fmt(start)} إلى {fmt(end)}')

emps = data.get('employees', [])
p_count = sum(1 for e in emps if e.get('status') == 'P')
v_count = sum(1 for e in emps if e.get('status') == 'V')
hosp = norm_filter(data.get('hospitalities', []), 'arrival')
prods = norm_filter(data.get('bakeryProductions', []), 'date')
ctr_sup = norm_filter(data.get('bakeryContractorSupplies', []), 'date')
meals = norm_filter(data.get('mealLogs', []), 'date')
maint = norm_filter(data.get('maintenanceRecords', []), 'date')
septic = norm_filter(data.get('septicRecords', []), 'date')
ts_batches = norm_filter(data.get('teaSugarBatches', []), 'date')
tea_sugar = norm_filter(data.get('teaSugarDisbursements', []), 'date')
incidents = norm_filter(filt(fetch_row_all('incident_reports', 'incident_reports'), 'incident_reports'), 'opened_at')
# Fetch meal waste entries (من ent:mealWaste في السحابة + الملفات المحلية)
mw_data = norm_filter(data.get('mealWaste', []), 'date')

def _week_days():
    return [(start + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]

def _sum_series(records, date_key, keys):
    m = {d: {k: 0 for k in keys} for d in _week_days()}
    for r in records:
        d = norm_date(r.get(date_key, ''))
        if d in m:
            for k in keys:
                m[d][k] += float(r.get(k, 0) or 0)
    return m

def _count_series(records, date_key):
    m = {d: 0 for d in _week_days()}
    for r in records:
        d = norm_date(r.get(date_key, ''))
        if d in m: m[d] += 1
    return m

def _est(series, day, key=None):
    if key is not None:
        series = {d: series[d][key] for d in series}
    vals = [series[d] for d in _week_days() if d in series and series[d]]
    if len(vals) < 1: return None
    idx = _week_days().index(day)
    nb = [series[_week_days()[di]] for di in (idx - 1, idx + 1) if 0 <= di < 7 and _week_days()[di] in series and series[_week_days()[di]]]
    if nb: return int(round(1.0 * sum(nb) / len(nb)))
    return int(round(1.0 * sum(vals) / len(vals)))

def _append_est_block(ws, rows):
    if not rows: return
    r = ws.max_row + 1
    for cells in rows:
        for ci, v in enumerate(cells, 1):
            ws.cell(row=r, column=ci, value=v)
        r += 1

today_str = datetime.now().strftime('%Y-%m-%d')
report_dir = 'C:\\Users\\Salem Magdy\\Desktop\\التقرير الاسبوعي للداتا'
os.makedirs(report_dir, exist_ok=True)
filename = os.path.join(report_dir, f'Lina_Weekly_{today_str}.xlsx')

base = filename.replace('.xlsx','')
i = 1
while os.path.exists(filename):
    filename = f'{base}_{i}.xlsx'
    i += 1

wb = Workbook()

ws = wb.active; ws.title = 'Summary'
style_sheet(ws, ['Item', 'Count'], [
    ['Total Employees (القوة)', len(emps)], ['Present (P)', p_count], ['Vacation (V)', v_count],
    ['Hospitality', len(hosp)], ['Bakery Production', len(prods)],
    ['Contractor Supply', len(ctr_sup)], ['Meals', len(meals)],
    ['Maintenance', len(maint)], ['Septic', len(septic)],
    ['Incidents', len(incidents)], ['Tea & Sugar (Batches)', len(ts_batches)], ['Tea & Sugar (Disbursed)', len(tea_sugar)],
    ['Meal Waste', len(mw_data)]
], [28, 12], title=f'Weekly Report {fmt(start)} to {fmt(end)}')

# Daily Statistics sheet
daily_stats = data.get('dailyStats', [])
daily_stats = norm_filter(daily_stats, 'date')
if daily_stats:
    ws_ds = wb.create_sheet('DailyStats')
    ds_by_date = {}
    for _s in sorted(daily_stats, key=lambda x: norm_date(x.get('date',''))):
        ds_by_date[norm_date(_s.get('date',''))] = _s
    _ds_keys = ['total', 'permP', 'permV', 'casP', 'hospGuests']
    kseries = {}
    for k in _ds_keys:
        kseries[k] = {d: (ds_by_date[d].get(k) or 0) for d in ds_by_date if (ds_by_date[d].get(k) or 0) != 0}
    _filled = set()
    for k in _ds_keys:
        for day in _week_days():
            rec = ds_by_date.get(day)
            if rec is not None and not (rec.get(k) or 0):
                e = _est(kseries[k], day)
                if e is not None:
                    rec[k] = e
                    _filled.add(day)
    ds_rows = []
    for day in _week_days():
        s = ds_by_date.get(day)
        if s is None:
            _tot = _est(kseries['total'], day) or 0
            if not _tot: continue
            estv = {k: (_est(kseries[k], day) or 0) for k in _ds_keys}
            _tot = estv['total']
            ds_rows.append([day, _tot,
                estv['permP'], f"{round(estv['permP'] * 100 / max(_tot,1))}%",
                estv['permV'], f"{round(estv['permV'] * 100 / max(_tot,1))}%",
                estv['casP'], f"{round(estv['casP'] * 100 / max(_tot,1))}%",
                estv['hospGuests']])
            _filled.add(day)
            continue
        _tot = s.get('total',0) or 0
        if day in _filled:
            ds_rows.append([day, _tot,
                s.get('permP',0), f"{round((s.get('permP',0) or 0) * 100 / max(_tot,1))}%",
                s.get('permV',0), f"{round((s.get('permV',0) or 0) * 100 / max(_tot,1))}%",
                s.get('casP',0), f"{round((s.get('casP',0) or 0) * 100 / max(_tot,1))}%",
                s.get('hospGuests',0)])
        else:
            ds_rows.append([day, _tot,
                s.get('permP',0), (str(s.get('permPPct',0)) + '%' if s.get('permPPct') is not None else '0%'),
                s.get('permV',0), (str(s.get('permVPct',0)) + '%' if s.get('permVPct') is not None else '0%'),
                s.get('casP',0), (str(s.get('casPPct',0)) + '%' if s.get('casPPct') is not None else '0%'),
                s.get('hospGuests',0)])
    style_sheet(ws_ds, ['Date', 'Total', 'Perm Present', 'Perm Present %', 'Perm Leave', 'Perm Leave %',
                         'Casual Present', 'Casual Present %', 'Guests'],
                ds_rows, [14,10,12,12,12,12,12,12,10], title='Daily Statistics')
    if len(ds_rows) > 1:
        data_end = 3 + len(ds_rows)
        avg = lambda idx: round(sum(r[idx] for r in ds_rows) / len(ds_rows))
        ws_ds.cell(row=data_end, column=1, value='معدل').font = Font(bold=True, size=11, color='1B5E20')
        ws_ds.cell(row=data_end, column=2, value=avg(1))
        ws_ds.cell(row=data_end, column=3, value=avg(2))
        ws_ds.cell(row=data_end, column=5, value=avg(4))
        ws_ds.cell(row=data_end, column=7, value=avg(6))
        ws_ds.cell(row=data_end, column=9, value=avg(8))

if prods:
    ws2 = wb.create_sheet('Bakery')
    _seen_prod = set()
    _prod_rows = []
    for p in sorted(prods, key=lambda x: norm_date(x.get('date',''))):
        _pk = norm_date(p.get('date','')) + '|' + str(p.get('breadCount','')) + '|' + str(p.get('flourUsed',''))
        if _pk in _seen_prod: continue
        _seen_prod.add(_pk)
        _prod_rows.append([norm_date(p.get('date','')), p.get('breadCount',0), p.get('flourUsed',0), p.get('branUsed',0), p.get('saltUsed',0), p.get('yeastUsed',0), p.get('dieselUsed',0)])
    style_sheet(ws2, ['Date', 'Bread', 'Flour', 'Bran', 'Salt', 'Yeast', 'Diesel'],
                _prod_rows, [14,16,10,10,10,10,10], title='Bakery Production')
    data_end = 3 + len(_prod_rows)
    ws2.cell(row=data_end, column=1, value='الإجمالي').font = Font(bold=True, size=11, color='1B5E20')
    ws2.cell(row=data_end, column=2, value=round(sum(r[1] for r in _prod_rows)))
    ws2.cell(row=data_end, column=3, value=round(sum(r[2] for r in _prod_rows),1))
    ws2.cell(row=data_end, column=4, value=round(sum(r[3] for r in _prod_rows),1))
    ws2.cell(row=data_end, column=5, value=round(sum(r[4] for r in _prod_rows),1))
    ws2.cell(row=data_end, column=6, value=round(sum(r[5] for r in _prod_rows),1))
    ws2.cell(row=data_end, column=7, value=round(sum(r[6] for r in _prod_rows),1))

if ctr_sup:
    ws3 = wb.create_sheet('Contractors')
    _seen_ctr = set()
    rows = []
    for c in sorted(ctr_sup, key=lambda x: norm_date(x.get('date',''))):
        _ck = norm_date(c.get('date','')) + '|' + str(c.get('name','')) + '|' + str(c.get('count','')) + '|' + str(c.get('price',''))
        if _ck in _seen_ctr: continue
        _seen_ctr.add(_ck)
        rows.append([norm_date(c.get('date','')), c.get('name',''), c.get('count',0), c.get('price',0),
                     int(c.get('count',0)or 0)*float(c.get('price',0)or 0)])
    style_sheet(ws3, ['Date', 'Name', 'Loaves', 'Price', 'Total'], rows, [14,20,10,10,12], title='Contractor Supply')
    data_end = 3 + len(rows)
    ws3.cell(row=data_end, column=1, value='الإجمالي').font = Font(bold=True, size=11, color='1B5E20')
    ws3.cell(row=data_end, column=3, value=sum(r[2] for r in rows))
    ws3.cell(row=data_end, column=5, value=round(sum(r[4] for r in rows),2))

if hosp:
    ws4 = wb.create_sheet('Hospitality')
    _seen_hosp = set()
    _hosp_rows = []
    for h in sorted(hosp, key=lambda x: norm_date(x.get('arrival',''))):
        _hk = f"{h.get('name','')}|{norm_date(h.get('arrival',''))}"
        if _hk in _seen_hosp: continue
        _seen_hosp.add(_hk)
        _hosp_rows.append([h.get('name',''), norm_date(h.get('arrival','')), norm_date(h.get('departure','')) if h.get('departure') else '', h.get('guests',1)])
    style_sheet(ws4, ['Name', 'Arrival', 'Departure', 'Guests'], _hosp_rows, [25,14,14,10], title='Hospitality')

if maint:
    ws5 = wb.create_sheet('Maintenance')
    _seen_mt = set()
    _mt_rows = []
    for m in sorted(maint, key=lambda x: norm_date(x.get('date',''))):
        _mk = f"{norm_date(m.get('date',''))}|{m.get('category','')}|{m.get('task','')}"
        if _mk in _seen_mt: continue
        _seen_mt.add(_mk)
        _mt_rows.append([norm_date(m.get('date','')), m.get('category',''), m.get('task',''), m.get('cost',0), m.get('responsible','')])
    style_sheet(ws5, ['Date', 'Category', 'Task', 'Cost', 'Responsible'], _mt_rows, [14,15,30,10,20], title='Maintenance')

if meals:
    ws6 = wb.create_sheet('Meals')
    _seen_ml = set()
    rows = []
    for m in sorted(meals, key=lambda x: norm_date(x.get('date',''))):
        _dl = norm_date(m.get('date',''))
        if not _dl or _dl in _seen_ml: continue
        _seen_ml.add(_dl)
        rows.append([_dl, m.get('breakfast',0), m.get('lunch',0), m.get('dinner',0),
                     int(m.get('breakfast',0)or 0)+int(m.get('lunch',0)or 0)+int(m.get('dinner',0)or 0)])
    style_sheet(ws6, ['Date', 'Breakfast', 'Lunch', 'Dinner', 'Total'], rows, [14,10,10,10,12], title='Meals')

if septic:
    ws7 = wb.create_sheet('Septic')
    _septic_rows = [{
        'date': norm_date(s.get('date','')),
        'name': s.get('name', s.get('sector','')),
        'trips': int(s.get('trips',0) or s.get('quantity',0) or 0),
        'qty': float(s.get('quantity', s.get('pumpQty', s.get('amount', s.get('حجم', (s.get('trips',0) or 0)*5)) )) or 0)
    } for s in septic if norm_date(s.get('date',''))]
    _septic_days = {d: 0 for d in _week_days()}
    _septic_total_by_day = {d: 0 for d in _week_days()}
    _septic_qty_by_day = {d: 0.0 for d in _week_days()}
    for r in _septic_rows:
        if r['date'] in _septic_days:
            _septic_days[r['date']] += 1
            _septic_total_by_day[r['date']] += r['trips']
            _septic_qty_by_day[r['date']] += r['qty']
    # متوسط النقلات اليومية في الأيام اللي فيها سجلات
    _dates_with = [d for d in _week_days() if _septic_total_by_day[d] > 0]
    _avg_trips = int(round(sum(_septic_total_by_day[d] for d in _dates_with) / max(len(_dates_with),1))) if _dates_with else 0
    _avg_qty = round(sum(_septic_qty_by_day[d] for d in _dates_with) / max(len(_dates_with),1), 1) if _dates_with else 0
    _septic_out = []
    for _day in _week_days():
        if _septic_total_by_day.get(_day, 0) > 0:
            for r in _septic_rows:
                if r['date'] == _day:
                    _septic_out.append([r['date'], r['name'], r['trips'], r['qty']])
    style_sheet(ws7, ['Date', 'Name', 'Trips', 'Quantity (m³)'], _septic_out, [14,30,10,12], title='Septic')

if incidents:
    ws8 = wb.create_sheet('Incidents')
    style_sheet(ws8, ['Date', 'Location', 'Category', 'Description', 'Status', 'Priority', 'Reporter'], [
        [(i.get('opened_at','') or i.get('date',''))[:10], i.get('location',''), i.get('type', i.get('category', '')), i.get('desc','') or i.get('description',''), i.get('status',''), i.get('priority',''), i.get('name','')]
        for i in sorted(incidents, key=lambda x: x.get('opened_at','') or x.get('date',''))
    ], [14,15,15,40,10,12,20], title='Incidents')

if tea_sugar:
    ws9 = wb.create_sheet('TeaSugar')
    style_sheet(ws9, ['Date', 'Code', 'Name', 'Tea', 'Sugar', 'Period'], [
        [norm_date(t.get('date','')), t.get('empCode',''), strip_emoji(t.get('empName',t.get('name',''))),
         t.get('teaPacks',0), t.get('sugarKg',0), t.get('period',t.get('type',''))]
        for t in sorted(tea_sugar, key=lambda x: norm_date(x.get('date','')))
    ], [14,12,25,10,10,20], title='Tea & Sugar')

if ts_batches:
    ws10 = wb.create_sheet('TeaSugarBatches')
    style_sheet(ws10, ['Date', 'Period', 'Tea Qty', 'Sugar Qty'], [
        [norm_date(b.get('date','')), b.get('period',''), b.get('teaQty',0), b.get('sugarQty',0)]
        for b in sorted(ts_batches, key=lambda x: norm_date(x.get('date','')))
    ], [14,20,10,10], title='Tea Sugar Batches')

if mw_data:
    # استبعاد السجلات التلقائية الفارغة (autoGenerated/autoFilled بكل القيم صفر) — ليست هدراً حقيقياً
    real_mw = [
        m for m in mw_data
        if not (m.get('autoGenerated') or m.get('autoFilled'))
        or float(m.get('wasteEng',0) or 0) + float(m.get('wasteWrk',0) or 0) + float(m.get('wasteGuests',0) or 0) + float(m.get('prepWaste',0) or 0) > 0
    ]
    ws11 = wb.create_sheet('MealWaste')
    _seen_mw = set()
    rows = []
    for m in sorted(real_mw, key=lambda x: (norm_date(x.get('date','')), str(x.get('meal','')))):
        _dkey = norm_date(m.get('date','')) + '|' + str(m.get('meal','')).strip()
        if _dkey in _seen_mw:
            continue
        _seen_mw.add(_dkey)
        waste = float(m.get('wasteEng',0)or 0) + float(m.get('wasteWrk',0)or 0) + float(m.get('wasteGuests',0)or 0) + float(m.get('prepWaste',0) or 0)
        ppl = (int(m.get('engAte',0)or 0) + int(m.get('wrkAte',0)or 0) + int(m.get('guests',0)or 0) + int(m.get('engTakeaway',0)or 0) + int(m.get('wrkTakeaway',0)or 0)) or 1
        cost = float(m.get('cost',0)or 0)
        wp_g = round(waste / ppl * 1000) if ppl > 0 else 0
        meals_str = str(m.get('meal',''))
        marker = ''
        if str(meals_str).endswith('(تلقائي)') or (m.get('autoFilled') and not (float(m.get('wasteEng',0)or 0) + float(m.get('wasteWrk',0)or 0) + float(m.get('wasteGuests',0)or 0)) ):
            marker = ' •'
        rows.append([norm_date(m.get('date','')), meals_str + marker, m.get('chef',m.get('responsible','')),
                      ppl, round(waste, 1), wp_g, round(cost)])
    real_real = [m for m in real_mw if not (m.get('autoGenerated') or m.get('autoFilled'))]
    style_sheet(ws11, ['Date', 'Meal', 'Chef', 'Meals Count', 'Waste (kg)', 'Waste/Person (g)', 'Cost (ج.م)'],
                rows, [14,10,20,12,12,16,12], title='Meal Waste')
    data_end = 3 + len(rows)
    ws11.cell(row=data_end, column=1, value='الإجمالي').font = Font(bold=True, size=11, color='1B5E20')
    ws11.cell(row=data_end, column=4, value=sum(r[3] for r in rows))
    ws11.cell(row=data_end, column=5, value=round(sum(r[4] for r in rows), 1))
    total_ppl = sum(r[3] for r in rows)
    ws11.cell(row=data_end, column=6, value=round(sum(r[4] for r in rows) / max(total_ppl, 1) * 1000))
    ws11.cell(row=data_end, column=7, value=round(sum(r[6] for r in rows)))

wb.save(filename)
print(f'[تم] {filename}')
