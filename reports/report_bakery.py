import requests, json, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

SUPABASE_URL = 'https://cwqghiqykohefaggedjl.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImN3cWdoaXF5a29oZWZhZ2dlZGpsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEwMjUyMjEsImV4cCI6MjA5NjYwMTIyMX0.3a3hRcNdmYQCtjYjBroAT6df1T_7oz-XWUeD3wagYw8'

def fetch_data():
    r = requests.get(f'{SUPABASE_URL}/rest/v1/sync_data?id=eq.alldata&select=data', headers={'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'})
    d = r.json()
    raw = d[0]['data']
    if isinstance(raw, str): raw = json.loads(raw)
    return raw

def make_bakery_report(data, months=3):
    prods = data.get('bakeryProductions', [])
    ctrs = data.get('bakeryContractorSupplies', [])
    cutoff = (datetime.now() - timedelta(days=30*months)).strftime('%Y-%m-%d')
    prods = [p for p in prods if p.get('date','') >= cutoff]
    ctrs = [c for c in ctrs if c.get('date','') >= cutoff]
    monthly = {}
    for p in prods:
        m = p['date'][:7]
        if m not in monthly: monthly[m] = {'flour':0,'bran':0,'salt':0,'yeast':0,'diesel':0,'bread_farm':0}
        monthly[m]['flour'] += float(p.get('flourUsed',0) or 0)
        monthly[m]['bran'] += float(p.get('branUsed',0) or 0)
        monthly[m]['salt'] += float(p.get('saltUsed',0) or 0)
        monthly[m]['yeast'] += float(p.get('yeastUsed',0) or 0)
        monthly[m]['diesel'] += float(p.get('dieselUsed',0) or 0)
        monthly[m]['bread_farm'] += int(p.get('breadCount',0) or 0)
    for c in ctrs:
        m = c['date'][:7]
        if m not in monthly: monthly[m] = {'flour':0,'bran':0,'salt':0,'yeast':0,'diesel':0,'bread_farm':0}
        monthly[m]['bread_farm'] = monthly[m].get('bread_farm',0) + int(c.get('count',0) or 0)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Monthly Bakery'
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill('solid', fgColor='1B5E20')
    thin = Side(style='thin')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    headers = ['الشهر', 'خبز الفرن', 'دقيق (كجم)', 'ردة (كجم)', 'ملح (كجم)', 'خميرة (كجم)', 'سولار (لتر)']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
        cell.border = border
    for r, m in enumerate(sorted(monthly.keys(), reverse=True), 2):
        d = monthly[m]
        vals = [m, d['bread_farm'], round(d['flour'],1), round(d['bran'],1), round(d['salt'],1), round(d['yeast'],1), round(d['diesel'],1)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border; cell.alignment = Alignment(horizontal='center')
    for c in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+c)].width = 16
    wb.save('C:\\Users\\Salem Magdy\\Desktop\\تقارير_الفرن.xlsx')
    print(f'[تم حفظ تقرير الفرن] {len(monthly)} شهر')

if __name__ == '__main__':
    data = fetch_data()
    make_bakery_report(data)
