import requests, json, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

SUPABASE_URL = 'https://cwqghiqykohefaggedjl.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImN3cWdoaXF5a29oZWZhZ2dlZGpsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEwMjUyMjEsImV4cCI6MjA5NjYwMTIyMX0.3a3hRcNdmYQCtjYjBroAT6df1T_7oz-XWUeD3wagYw8'

def fetch():
    r = requests.get(f'{SUPABASE_URL}/rest/v1/sync_data?id=eq.alldata&select=data', headers={'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'})
    raw = r.json()[0]['data']
    return json.loads(raw) if isinstance(raw, str) else raw

def style_sheet(ws, headers, rows, col_widths=None):
    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill('solid', fgColor='1B5E20')
    thin = Side(style='thin'); border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = border
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border; cell.alignment = Alignment(horizontal='center')
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

print('[جاري سحب البيانات...]')
data = fetch()
wb = Workbook()

# 1. الموظفين
emps = data.get('employees', [])
ws = wb.active; ws.title = 'الموظفين'
style_sheet(ws, ['الكود','الاسم','الحالة','التعاقد','الإدارة','المسمى','المحافظة','المبنى','الغرفة','تاريخ التعيين'],
  [[e.get('code',''), e.get('name',''), e.get('status',''), e.get('contract',''),
    e.get('dept',''), e.get('title',''), e.get('gov',''), e.get('sector',''), e.get('room',''), e.get('hireDate','')] for e in emps],
  [12,25,8,10,20,20,15,12,10,14])

# 2. الإجازات
vacs = data.get('vacations', [])
if vacs:
    ws = wb.create_sheet('الإجازات')
    style_sheet(ws, ['الموظف','الكود','من','إلى','نوع'],
      [[v.get('employeeName',v.get('name','')), v.get('code',v.get('employeeCode','')),
        v.get('start',v.get('startDate','')), v.get('end',v.get('endDate','')), v.get('type','')] for v in vacs],
      [25,12,14,14,15])

# 3. الضيافة
hosp = data.get('hospitalities', [])
if hosp:
    ws = wb.create_sheet('الضيافة')
    style_sheet(ws, ['الاسم','الوصول','المغادرة','العدد','النوع'],
      [[h.get('name',''), h.get('arrival',''), h.get('departure',''), h.get('guests',1), h.get('type','')] for h in hosp],
      [25,14,14,8,15])

# 4. الصيانة
maint = data.get('maintenanceRecords', [])
if maint:
    ws = wb.create_sheet('الصيانة')
    style_sheet(ws, ['التاريخ','التصنيف','المهمة','التكلفة','المسؤول'],
      [[m.get('date',''), m.get('category',''), m.get('task',''), m.get('cost',0), m.get('responsible','')] for m in maint],
      [14,15,30,10,20])

# 5. البيارات
septic = data.get('septicRecords', [])
if septic:
    ws = wb.create_sheet('البيارات')
    style_sheet(ws, ['التاريخ','الاسم','عدد الرحلات','الكمية'],
      [[s.get('date',''), s.get('name',s.get('sector','')), s.get('trips',0), s.get('quantity',0)] for s in septic],
      [14,20,12,10])

# 6. المخزن - الأصناف
items = data.get('inventoryItems', [])
if items:
    ws = wb.create_sheet('أصناف المخزن')
    style_sheet(ws, ['الاسم','الكمية','الوحدة'],
      [[i.get('name',''), i.get('qty',i.get('quantity',0)), i.get('unit','')] for i in items],
      [25,10,10])

# 7. المخزن - بونات الصرف
vouch = data.get('inventoryVouchers', [])
if vouch:
    ws = wb.create_sheet('بونات الصرف')
    style_sheet(ws, ['التاريخ','الصنف','الكمية','الجهة'],
      [[v.get('date',''), v.get('itemName',v.get('item','')), v.get('qty',0), v.get('department',v.get('entity',''))] for v in vouch],
      [14,25,10,20])

# 8. شاي وسكر
ts = data.get('teaSugarDisbursements', [])
if ts:
    ws = wb.create_sheet('شاي وسكر')
    style_sheet(ws, ['التاريخ','الموظف','شاي','سكر','الفترة'],
      [[t.get('date',''), t.get('empCode',t.get('name','')), t.get('teaPacks',0), t.get('sugarKg',0), t.get('period',t.get('type',''))] for t in ts],
      [14,20,8,8,12])

# 9. الوجبات
meals = data.get('mealLogs', [])
if meals:
    ws = wb.create_sheet('الوجبات')
    style_sheet(ws, ['التاريخ','الفطور','الغداء','العشاء','الإجمالي'],
      [[m.get('date',''), m.get('breakfast',0), m.get('lunch',0), m.get('dinner',0),
        (int(m.get('breakfast',0) or 0) + int(m.get('lunch',0) or 0) + int(m.get('dinner',0) or 0))] for m in meals],
      [14,10,10,10,12])

# 10. المقاولين
ctrs = data.get('contractors', [])
if ctrs:
    ws = wb.create_sheet('المقاولين')
    style_sheet(ws, ['الاسم','القطاع','الغرفة','بداية العقد','نهاية العقد'],
      [[c.get('name',''), c.get('sector',''), c.get('room',''), c.get('startDate',''), c.get('endDate','')] for c in ctrs],
      [25,15,10,14,14])

# 11. إنتاج الفرن
prods = data.get('bakeryProductions', [])
if prods:
    ws = wb.create_sheet('إنتاج الفرن')
    style_sheet(ws, ['التاريخ','عدد الأرغفة','دقيق','ردة','ملح','خميرة','سولار'],
      [[p.get('date',''), p.get('breadCount',0), p.get('flourUsed',0), p.get('branUsed',0),
        p.get('saltUsed',0), p.get('yeastUsed',0), p.get('dieselUsed',0)] for p in prods],
      [14,12,10,10,10,10,10])

# 12. توريد مقاولين
ctr_sup = data.get('bakeryContractorSupplies', [])
if ctr_sup:
    ws = wb.create_sheet('توريد مقاولين')
    style_sheet(ws, ['التاريخ','المقاول','عدد الأرغفة','السعر','الإجمالي'],
      [[c.get('date',''), c.get('name',''), c.get('count',0), c.get('price',0),
        (int(c.get('count',0) or 0) * float(c.get('price',0) or 0))] for c in ctr_sup],
      [14,20,12,10,12])

# 13. فواتير المخبز
inv = data.get('bakeryInvoices', [])
if inv:
    ws = wb.create_sheet('فواتير المخبز')
    style_sheet(ws, ['التاريخ','البيان','المبلغ'],
      [[i.get('date',''), i.get('description',i.get('item','')), i.get('amount',0)] for i in inv],
      [14,30,12])

# 14. الصيانة الدورية
pm = data.get('periodicMaintenance', [])
if pm:
    ws = wb.create_sheet('صيانة دورية')
    style_sheet(ws, ['المهمة','التكرار','آخر تنفيذ','المسؤول'],
      [[p.get('name',p.get('task','')), p.get('freq',p.get('frequency','')), p.get('lastDone',p.get('lastDate','')), p.get('responsible','')] for p in pm],
      [30,12,14,20])

# 15. بلاغات الأعطال
reports = data.get('incident_reports', [])
if reports:
    ws = wb.create_sheet('بلاغات الأعطال')
    style_sheet(ws, ['التاريخ','الموقع','التصنيف','الوصف','الحالة'],
      [[r.get('date',''), r.get('location',''), r.get('category',''), r.get('description',''), r.get('status','')] for r in reports],
      [14,15,15,40,10])

path = 'C:\\Users\\Salem Magdy\\Desktop\\تقرير_لينة_الشامل.xlsx'
wb.save(path)
print(f'[تم] تقرير شامل — {len(emps)} موظف، {len(prods)} إنتاج، {len(maint)} صيانة، {len(ctr_sup)} توريد، {len(hosp)} ضيافة')
