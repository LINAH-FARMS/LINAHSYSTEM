import openpyxl, json, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('C:/Users/Salem Magdy/Downloads/الإدارة الفنية.xlsx')
ws = wb.active
print(f"Sheet: {ws.title}")
print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")

# Print first 10 rows
for r in range(1, min(12, ws.max_row + 1)):
    row_data = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(r, c).value
        row_data.append(str(val).strip() if val is not None else "")
    print(f"Row {r}: {' | '.join(row_data)}")

print("\n--- All data ---")
for r in range(2, ws.max_row + 1):
    row_data = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(r, c).value
        row_data.append(str(val).strip() if val is not None else "")
    if any(row_data):
        print(f"  {' | '.join(row_data)}")
