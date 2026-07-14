import openpyxl, json, sys, urllib.request
sys.stdout.reconfigure(encoding='utf-8')

apiKey = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImN3cWdoaXF5a29oZWZhZ2dlZGpsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEwMjUyMjEsImV4cCI6MjA5NjYwMTIyMX0.3a3hRcNdmYQCtjYjBroAT6df1T_7oz-XWUeD3wagYw8'

# Read Supabase
req = urllib.request.Request('https://cwqghiqykohefaggedjl.supabase.co/rest/v1/sync_data?id=eq.alldata&select=data',
    headers={'apikey': apiKey, 'Authorization': f'Bearer {apiKey}'})
with urllib.request.urlopen(req, timeout=15) as r:
    result = json.loads(r.read())
data = result[0]['data']
if isinstance(data, str): data = json.loads(data)
emps = data['employees']
rc = data['roomsCapacity']
sectors_in_rc = set(x['sector'] for x in rc)

# ===== FILE 1: بيانات سكن العاملين 2026 (1) =====
print("=" * 80)
print("FILE 1: بيانات سكن العاملين للعام 2026 (1)")
print("=" * 80)
wb1 = openpyxl.load_workbook(r'C:\Users\Salem Magdy\Downloads\بيان سكن العاملين للعام 2026 (1).xlsx')
ws1 = wb1.active
print(f"Sheet: {ws1.title}, Rows: {ws1.max_row}, Cols: {ws1.max_column}")
# Print headers
hdr1 = []
for c in range(1, ws1.max_column + 1):
    hdr1.append(str(ws1.cell(1, c).value or '').strip())
print(f"Headers: {hdr1}")

file1_emps = []
for r in range(2, ws1.max_row + 1):
    name = str(ws1.cell(r, 1).value or '').strip()
    code = str(ws1.cell(r, 2).value or '').strip()
    loc = str(ws1.cell(r, 3).value or '').strip()
    if name and code:
        file1_emps.append({'name': name, 'code': code, 'loc': loc})

print(f"Employees in file: {len(file1_emps)}")

# ===== FILE 2: الإدارة الفنية (1) =====
print("\n" + "=" * 80)
print("FILE 2: الإدارة الفنية (1)")
print("=" * 80)
wb2 = openpyxl.load_workbook(r'C:\Users\Salem Magdy\Downloads\الإدارة الفنية (1).xlsx')
ws2 = wb2.active
print(f"Sheet: {ws2.title}, Rows: {ws2.max_row}, Cols: {ws2.max_column}")
hdr2 = []
for c in range(1, ws2.max_column + 1):
    hdr2.append(str(ws2.cell(1, c).value or '').strip())
print(f"Headers: {hdr2}")

file2_emps = []
for r in range(2, ws2.max_row + 1):
    name = str(ws2.cell(r, 1).value or '').strip()
    code = str(ws2.cell(r, 2).value or '').strip()
    loc = str(ws2.cell(r, 3).value or '').strip()
    if name and code:
        file2_emps.append({'name': name, 'code': code, 'loc': loc})

print(f"Employees in file: {len(file2_emps)}")

# ===== MAP location from files to sector+room =====
# File 1 mapping (same as before)
file1_map = {
    'الجزورين غرفة 1': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 1'),
    'الجزورين غ 1': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 1'),
    'الجزورين غ 2': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 2'),
    'الجزورين غ 3': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 3'),
    'الجزورين غ 4': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 4'),
    'الجزورين غ 5': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 5'),
    'الجزورين غ 6': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 6'),
    'الجزورين غ 7': ('سكن العاملين (سكن الجيزوارين)', 'الجيم غرفه 7'),
    'الجزورين غ 8': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 8'),
    'الجزورين غ 9': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 9'),
    'السكن الجديد F3': ('سكن العاملين الجديد 2025 (F)', 'غرفه F3'),
    'السكن الجديد F6': ('سكن العاملين الجديد 2025 (F)', 'غرفه F6'),
    'النخالين غ 1': ('سكن العاملين (سكن النخالين)', 'غرفه نخاليين 1'),
    'النخالين غ 2': ('سكن العاملين (سكن النخالين)', 'غرفه نخاليين 2'),
    'ق 17': ('سكن القطاعات', 'قطاع رقم ( 17 )'),
    'قطاع 1': ('سكن القطاعات', 'قطاع رقم ( 1 )'),
    'ق 1': ('سكن القطاعات', 'قطاع رقم ( 1 )'),
    'ق 3': ('سكن القطاعات', 'قطاع رقم ( 3 )'),
    'قطاع 3': ('سكن القطاعات', 'قطاع رقم ( 3 )'),
    'ق 6': ('سكن القطاعات', 'قطاع رقم ( 6 )'),
    'ق 21': ('سكن القطاعات', 'قطاع رقم ( 21 )'),
    'ق 24': ('سكن القطاعات', 'قطاع رقم ( 24 )'),
    'ق 25': ('سكن القطاعات', 'قطاع رقم ( 25 )'),
    'ق 26': ('سكن القطاعات', 'قطاع رقم ( 26 )'),
    'ق 27': ('سكن القطاعات', 'قطاع رقم ( 27 )'),
    'ق 28': ('سكن القطاعات', 'قطاع رقم ( 28 )'),
    'ق 29': ('سكن القطاعات', 'قطاع رقم ( 29 )'),
    'ق 30': ('سكن القطاعات', 'قطاع رقم ( 30 )'),
    'ق 32': ('سكن القطاعات', 'قطاع رقم ( 32 )'),
    'ق 33': ('سكن القطاعات', 'قطاع رقم ( 33 )'),
}

# File 2 mapping - same structure as file 1 (الإدارة الفنية is a subset)
file2_map = file1_map.copy()

# ===== Compare both files against Supabase =====
def compare_file(file_emps, mapping, label):
    correct = 0
    wrong = 0
    not_in_program = 0
    not_mapped = 0
    wrong_list = []
    not_found_list = []
    
    for item in file_emps:
        loc = item['loc']
        if loc not in mapping:
            not_mapped += 1
            continue
        target_sector, target_room = mapping[loc]
        emp = next((e for e in emps if str(e.get('code','')) == item['code']), None)
        if not emp:
            not_in_program += 1
            not_found_list.append(f"  {item['code']:>6} | {item['name'][:35]:<35} | {loc}")
            continue
        actual_sector = emp.get('sector','')
        actual_room = emp.get('room','')
        if actual_sector == target_sector and actual_room == target_room:
            correct += 1
        else:
            wrong += 1
            wrong_list.append(f"  {item['code']:>6} | {item['name'][:35]:<35} | Expected: {target_sector} / {target_room}")
            wrong_list.append(f"         | {'':35} | Actual  : {actual_sector} / {actual_room}")
    
    print(f"\n--- {label} ---")
    print(f"  Total employees: {len(file_emps)}")
    print(f"  Correct: {correct}")
    print(f"  Wrong: {wrong}")
    print(f"  Not in program: {not_in_program}")
    print(f"  Location not mapped: {not_mapped}")
    if wrong_list:
        print(f"\n  ❌ WRONG ({wrong}):")
        for w in wrong_list:
            print(w)
    if not_found_list:
        print(f"\n  ⚠️ NOT IN PROGRAM ({not_in_program}):")
        for n in not_found_list:
            print(n)
    
    return correct, wrong, not_in_program, not_mapped

compare_file(file1_emps, file1_map, "بيان سكن العاملين 2026 (1)")
compare_file(file2_emps, file2_map, "الإدارة الفنية (1)")
