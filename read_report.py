import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r'C:\Users\Salem Magdy\Desktop\تقرير_السكن.xlsx')
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== {sn} (rows={ws.max_row}, cols={ws.max_column}) ===")
    for r in range(1, min(ws.max_row + 1, 120)):
        vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            vals.append(str(v).strip() if v is not None else '')
        print(f"  {' | '.join(vals)}")
