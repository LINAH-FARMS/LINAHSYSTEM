import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

# FILE 1
print("=" * 80)
print("FILE 1: بيانات سكن العاملين للعام 2026 (1)")
print("=" * 80)
wb1 = openpyxl.load_workbook(r'C:\Users\Salem Magdy\Downloads\بيان سكن العاملين للعام 2026 (1).xlsx')
ws1 = wb1.active
# Print first 10 rows to understand structure
for r in range(1, min(15, ws1.max_row + 1)):
    vals = []
    for c in range(1, ws1.max_column + 1):
        v = ws1.cell(r, c).value
        vals.append(str(v).strip() if v is not None else '')
    print(f"Row {r}: {' | '.join(vals)}")

# FILE 2
print("\n" + "=" * 80)
print("FILE 2: الإدارة الفنية (1)")
print("=" * 80)
wb2 = openpyxl.load_workbook(r'C:\Users\Salem Magdy\Downloads\الإدارة الفنية (1).xlsx')
ws2 = wb2.active
for r in range(1, min(15, ws2.max_row + 1)):
    vals = []
    for c in range(1, ws2.max_column + 1):
        v = ws2.cell(r, c).value
        vals.append(str(v).strip() if v is not None else '')
    print(f"Row {r}: {' | '.join(vals)}")
