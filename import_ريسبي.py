#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ريسبي للاستيراد — يقرأ ملف الريسبي.xlsx ويولّد JSON للبرنامج
"""
import openpyxl, json, sys, os, re
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_FILE = r"C:\Users\Salem Magdy\Desktop\الريسبي.xlsx"
OUTPUT_FILE = r"C:\Users\Salem Magdy\Desktop\LINAHSYSTEM\recipes_import.json"

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# ===== قراءة الوصفات =====
recipes = []
for sn in wb.sheetnames:
    if sn == 'اسعار':
        continue
    ws = wb[sn]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if not rows:
        continue
    header = rows[0]
    # اول خليه = اسم الوجبه, التاني = "لعدد 380 موظف"
    recipe_name = str(header[0]).strip() if header[0] else sn.strip()
    # استخراج عدد الافراد من الخانه التانيه
    base = 380
    if len(header) > 1 and header[1]:
        m = re.search(r'(\d+)', str(header[1]))
        if m:
            base = int(m.group(1))
    items = []
    for row in rows[1:]:
        qty = row[0]
        unit = str(row[1]).strip() if row[1] else ''
        ing_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        if not ing_name or qty is None:
            continue
        try:
            qty = float(qty)
        except:
            continue
        items.append({"q": qty, "u": unit, "n": ing_name})
    if items:
        recipes.append({"name": recipe_name, "base": base, "items": items})

# ===== قراءة الأسعار =====
ws_p = wb['اسعار']
prices = {}
for row in ws_p.iter_rows(min_row=2, max_row=ws_p.max_row, values_only=True):
    name = str(row[1]).strip() if row[1] else ''
    unit = str(row[2]).strip() if row[2] else ''
    price = row[3]
    if name and price is not None:
        try:
            prices[f"{name}|{unit}"] = float(price)
        except:
            pass

# ===== كتابة JSON =====
output = {
    "recipes": recipes,
    "prices": prices,
    "generated_from": os.path.basename(EXCEL_FILE),
    "generated_at": str(__import__('datetime').datetime.now())
}

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"✅ تم تصدير {len(recipes)} وصفة و {len(prices)} سعر إلى:")
print(f"   {OUTPUT_FILE}")
