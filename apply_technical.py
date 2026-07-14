import openpyxl, json, sys, urllib.request, re
sys.stdout.reconfigure(encoding='utf-8')

apiKey = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImN3cWdoaXF5a29oZWZhZ2dlZGpsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEwMjUyMjEsImV4cCI6MjA5NjYwMTIyMX0.3a3hRcNdmYQCtjYjBroAT6df1T_7oz-XWUeD3wagYw8'

def get_supabase():
    req = urllib.request.Request(
        'https://cwqghiqykohefaggedjl.supabase.co/rest/v1/sync_data?id=eq.alldata&select=data',
        headers={'apikey': apiKey, 'Authorization': f'Bearer {apiKey}'}
    )
    with urllib.request.urlopen(req, timeout=15) as r:
        result = json.loads(r.read())
    data = result[0]['data']
    if isinstance(data, str): data = json.loads(data)
    return data

def save_supabase(data):
    import datetime
    ts = datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.') + f'{datetime.datetime.now(datetime.timezone.utc).microsecond // 1000:03d}+00:00'
    payload = json.dumps({'data': data, 'updated_at': ts})
    req = urllib.request.Request(
        'https://cwqghiqykohefaggedjl.supabase.co/rest/v1/sync_data?id=eq.alldata',
        data=payload.encode('utf-8'),
        headers={'apikey': apiKey, 'Authorization': f'Bearer {apiKey}', 'Content-Type': 'application/json', 'Prefer': 'return=minimal'},
        method='PATCH'
    )
    with urllib.request.urlopen(req, timeout=15) as r:
        return r.status

# Read file
wb = openpyxl.load_workbook('C:/Users/Salem Magdy/Downloads/الإدارة الفنية.xlsx')
ws = wb.active
rows_data = []
for r in range(2, ws.max_row + 1):
    name = str(ws.cell(r, 1).value or '').strip()
    code = str(ws.cell(r, 2).value or '').strip()
    loc = str(ws.cell(r, 3).value or '').strip()
    if name and loc:
        rows_data.append({'name': name, 'code': code, 'loc': loc})

# MAPPING
sector_room_map = {
    # الجزورين = سكن الجيزوارين
    'الجزورين غرفة 1': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 1'),
    'الجزورين غ 1': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 1'),
    'الجزورين غ 2': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 2'),
    'الجزورين غ 3': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 3'),
    'الجزورين غ 4': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 4'),
    'الجزورين غ 5': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 5'),
    'الجزورين غ 6': ('سكن العاملين (سكن الجيزوارين)', 'غرفه 6'),
    # السكن الجديد F
    'السكن الجديد F3': ('سكن العاملين الجديد 2025 (F)', 'غرفه F3'),
    'السكن الجديد F6': ('سكن العاملين الجديد 2025 (F)', 'غرفه F6'),
    # النخالين
    'النخالين غ 1': ('سكن العاملين (سكن النخالين)', 'غرفه نخاليين 1'),
    'النخالين غ 2': ('سكن العاملين (سكن النخالين)', 'غرفه نخاليين 2'),
    # قطاعات
    'ق 17': ('سكن القطاعات', 'قطاع رقم ( 17 )'),
    'قطاع 1': ('سكن القطاعات', 'قطاع رقم ( 1 )'),
    'ق 1': ('سكن القطاعات', 'قطاع رقم ( 1 )'),
    'ق 3': ('سكن القطاعات', 'قطاع رقم ( 3 )'),
    'قطاع 3': ('سكن القطاعات', 'قطاع رقم ( 3 )'),
    'ق 6': ('سكن القطاعات', 'قطاع رقم ( 6 )'),
    'ق 21': ('سكن القطاعات', 'قطاع رقم ( 21 )'),
    'ق 24': ('سكن القطاعات', 'قطاع رقم ( 24 )'),
    'ق 25': ('سكن القطاعات', 'قطاع رقم ( 25 )'),
    'ق 26': ('سكن القطاعات', 'قطاع رقم ( 26 )'),
    'ق 27': ('سكن القطاعات', 'قطاع رقم ( 27 )'),
    'ق 28': ('سكن القطاعات', 'قطاع رقم ( 28 )'),
    'ق 29': ('سكن القطاعات', 'قطاع رقم ( 29 )'),
    'ق 30': ('سكن القطاعات', 'قطاع رقم ( 30 )'),
    'ق 32': ('سكن القطاعات', 'قطاع رقم ( 32 )'),
    'ق 33': ('سكن القطاعات', 'قطاع رقم ( 33 )'),
}

data = get_supabase()
employees = data['employees']

matched = 0
not_found = 0
moved = 0
already = 0
unclear = 0
not_found_list = []
moved_list = []

for item in rows_data:
    loc = item['loc']
    if loc not in sector_room_map:
        unclear += 1
        continue
    target_sector, target_room = sector_room_map[loc]
    # Find employee
    emp = None
    if item['code']:
        for e in employees:
            if str(e.get('code', '')) == item['code']:
                emp = e; break
    if not emp:
        for e in employees:
            if e.get('name', '') == item['name']:
                emp = e; break
    if not emp:
        not_found += 1
        not_found_list.append(f"{item['code']} | {item['name']} | {loc}")
        continue
    matched += 1
    old_sector = emp.get('sector', '')
    old_room = emp.get('room', '')
    if old_sector == target_sector and old_room == target_room:
        already += 1
    else:
        emp['sector'] = target_sector
        emp['room'] = target_room
        moved += 1
        moved_list.append(f"  {item['code']:>5} | {item['name'][:30]:<30} | {old_sector} - {old_room} -> {target_sector} - {target_room}")

print(f"Total in file: {len(rows_data)}")
print(f"Clear mapping: {len(rows_data) - unclear}")
print(f"Matched: {matched}")
print(f"  Already correct: {already}")
print(f"  Moved: {moved}")
print(f"Not found: {not_found}")
print(f"Unclear (skipped): {unclear}")

if moved_list:
    print(f"\n=== MOVED ({moved}) ===")
    for m in moved_list: print(m)

if not_found_list:
    print(f"\n=== NOT FOUND ({not_found}) ===")
    for n in not_found_list: print(n)

if moved > 0:
    data['employees'] = employees
    status = save_supabase(data)
    print(f"\nSaved to Supabase: {status}")
else:
    print("\nNothing to save")
