import openpyxl, json, sys
sys.stdout.reconfigure(encoding='utf-8')

out_path = r'C:\Users\SALEMM~1\AppData\Local\Temp\excel_data.txt'

wb = openpyxl.load_workbook(r'C:\Users\Salem Magdy\Desktop\المخازن.xlsx')
lines = []
lines.append('عدد الشيتات: ' + str(len(wb.sheetnames)))
for name in wb.sheetnames:
    ws = wb[name]
    lines.append('\n=== شيت: ' + name + ' ===')
    lines.append('الصفوف: ' + str(ws.max_row) + ', الأعمدة: ' + str(ws.max_column))
    # Print headers
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value or ''))
    lines.append(' | '.join(headers))
    # Print all data rows as JSON-like structure
    rows_data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        row_vals = [str(v or '') for v in row]
        rows_data.append(row_vals)
    lines.append('عدد الصفوف: ' + str(len(rows_data)))
    # Print first 3 rows
    for r in rows_data[:3]:
        lines.append(' | '.join(r))
    if len(rows_data) > 3:
        lines.append('...')
    # Save full data as JSON
    json_path = r'C:\Users\SALEMM~1\AppData\Local\Temp\excel_' + name + '.json'
    with open(json_path, 'w', encoding='utf-8') as jf:
        json.dump({'headers': headers, 'rows': rows_data}, jf, ensure_ascii=False)
    lines.append('محفوظ في: ' + json_path)

with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print('OK: ' + out_path)
