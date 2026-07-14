import json, sys, urllib.request
sys.stdout.reconfigure(encoding='utf-8')
apiKey = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImN3cWdoaXF5a29oZWZhZ2dlZGpsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEwMjUyMjEsImV4cCI6MjA5NjYwMTIyMX0.3a3hRcNdmYQCtjYjBroAT6df1T_7oz-XWUeD3wagYw8'

req = urllib.request.Request('https://cwqghiqykohefaggedjl.supabase.co/rest/v1/sync_data?id=eq.alldata&select=data',
    headers={'apikey': apiKey, 'Authorization': f'Bearer {apiKey}'})
with urllib.request.urlopen(req, timeout=15) as r:
    result = json.loads(r.read())
data = result[0]['data']
if isinstance(data, str): data = json.loads(data)

emps = data['employees']
rc = data['roomsCapacity']

# Build sector -> room -> [employees]
sectors = {}
for r in rc:
    s = r['sector']
    if s not in sectors: sectors[s] = {}
    sectors[s][r['number']] = {'beds': r['beds'], 'emps': []}

for e in emps:
    s = e.get('sector', '')
    rm = e.get('room', '')
    if s in sectors and rm in sectors[s]:
        sectors[s][rm]['emps'].append(e)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
wb = openpyxl.Workbook()

# Sheet 1: By sector/room
ws1 = wb.active
ws1.title = 'السكن حسب القطاعات'
ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 10

header_font = Font(bold=True, size=12)
header_fill = PatternFill('solid', fgColor='1565c0')
header_font_w = Font(bold=True, size=12, color='ffffff')
header_align = Alignment(horizontal='center', vertical='center')
thin = Side(style='thin')
border = Border(top=thin, left=thin, right=thin, bottom=thin)

ws1.append(['الاسم', 'الكود', 'الحالة', 'القطاع', 'الغرفة', 'السعة'])
for c in range(1, 7):
    ws1.cell(1, c).font = header_font_w
    ws1.cell(1, c).fill = header_fill
    ws1.cell(1, c).alignment = header_align

row = 2
for sname in sorted(sectors.keys()):
    sector_start = row
    for rname in sorted(sectors[sname].keys()):
        room = sectors[sname][rname]
        if not room['emps']:
            ws1.append(['', '', '', sname, rname, room['beds']])
            for c in range(1, 7): ws1.cell(row, c).border = border
            row += 1
        else:
            for e in room['emps']:
                ws1.append([e.get('name',''), e.get('code',''), e.get('status',''), sname, rname, room['beds']])
                for c in range(1, 7): ws1.cell(row, c).border = border
                row += 1
    # Merge sector cell
    if row - sector_start > 1:
        ws1.merge_cells(start_row=sector_start, start_column=4, end_row=row-1, end_column=4)

# Sheet 2: Employees without valid sector/room
ws2 = wb.create_sheet('موظفين بدون سكن')
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 15
ws2.append(['الاسم', 'الكود', 'الحالة', 'القطاع المسجل', 'الغرفة المسجلة'])
for c in range(1, 6):
    ws2.cell(1, c).font = header_font_w
    ws2.cell(1, c).fill = header_fill
    ws2.cell(1, c).alignment = header_align

valid_sectors = set(r['sector'] for r in rc)
valid_rooms = {}
for r in rc:
    if r['sector'] not in valid_rooms: valid_rooms[r['sector']] = set()
    valid_rooms[r['sector']].add(r['number'])

missing = [e for e in emps if e.get('sector') not in valid_sectors or e.get('sector') not in valid_rooms or e.get('room') not in valid_rooms.get(e.get('sector',''), set())]
# Also employees without sector/room at all
missing2 = [e for e in emps if not e.get('sector') or not e.get('room')]
all_missing = []
seen = set()
for e in missing + missing2:
    if e.get('code') not in seen or e.get('name') not in seen:
        all_missing.append(e)
        seen.add(e.get('code'))
        seen.add(e.get('name'))

for e in all_missing:
    ws2.append([e.get('name',''), e.get('code',''), e.get('status',''), e.get('sector',''), e.get('room','')])

# Sheet 3: Statistics
ws3 = wb.create_sheet('إحصائيات')
ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 20
ws3.append(['الإحصائية', 'القيمة'])
ws3.cell(1, 1).font = header_font_w; ws3.cell(1, 1).fill = header_fill
ws3.cell(1, 2).font = header_font_w; ws3.cell(1, 2).fill = header_fill

total_emps = len(emps)
total_beds = sum(int(r.get('beds',0)) for r in rc)
occupied = sum(1 for e in emps if e.get('status') == 'P' and e.get('sector') in valid_sectors and e.get('room'))
on_vacation = sum(1 for e in emps if e.get('status') == 'V' and e.get('sector') in valid_sectors and e.get('room'))
with_sector = sum(1 for e in emps if e.get('sector') in valid_sectors)
sector_count = len(valid_sectors)
room_count = len(rc)

ws3.append(['إجمالي الموظفين', total_emps])
ws3.append(['موظفين حاضرين', occupied])
ws3.append(['موظفين في إجازة', on_vacation])
ws3.append(['موظفين في سكن', occupied + on_vacation])
ws3.append(['موظفين بدون سكن', total_emps - with_sector])
ws3.append(['عدد القطاعات', sector_count])
ws3.append(['عدد الغرف', room_count])
ws3.append(['إجمالي السعة', total_beds])
ws3.append(['المشغول (حاضر+إجازة)', occupied + on_vacation])
ws3.append(['الفاضي', total_beds - (occupied + on_vacation)])

# Sheet 4: Per-sector summary
ws4 = wb.create_sheet('ملخص القطاعات')
ws4.column_dimensions['A'].width = 40
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 12
ws4.column_dimensions['E'].width = 12
ws4.append(['القطاع', 'عدد الغرف', 'السعة', 'مشغول', 'فاضي'])
for c in range(1, 6):
    ws4.cell(1, c).font = header_font_w
    ws4.cell(1, c).fill = header_fill
    ws4.cell(1, c).alignment = header_align

for sname in sorted(sectors.keys()):
    s_rooms = sectors[sname]
    rcount = len(s_rooms)
    sbeds = sum(int(s_rooms[rn]['beds']) for rn in s_rooms)
    socc = sum(len(s_rooms[rn]['emps']) for rn in s_rooms)
    ws4.append([sname, rcount, sbeds, socc, sbeds - socc])

outpath = r'C:\Users\Salem Magdy\Desktop\تقرير_السكن.xlsx'
wb.save(outpath)
print(f"✅ تم حفظ التقرير: {outpath}")
