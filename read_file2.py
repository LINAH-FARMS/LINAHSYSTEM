import openpyxl, json, sys, urllib.request
sys.stdout.reconfigure(encoding='utf-8')

apiKey = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImN3cWdoaXF5a29oZWZhZ2dlZGpsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODEwMjUyMjEsImV4cCI6MjA5NjYwMTIyMX0.3a3hRcNdmYQCtjYjBroAT6df1T_7oz-XWUeD3wagYw8'

# Read Supabase
req = urllib.request.Request('https://cwqghiqykohefaggedjl.supabase.co/rest/v1/sync_data?id=eq.alldata&select=data',
    headers={'apikey': apiKey, 'Authorization': f'Bearer {apiKey}'})
with urllib.request.urlopen(req, timeout=15) as r:
    result = json.loads(r.read())
data = result[0]['data']
if isinstance(data, str): data = json.loads(data)
emps = data['employees']
rc = data['roomsCapacity']
rc_lookup = {}
for r in rc:
    if r['sector'] not in rc_lookup: rc_lookup[r['sector']] = {}
    rc_lookup[r['sector']][r['number']] = r['beds']

# ===== FILE 2: الإدارة الفنية (1) =====
wb2 = openpyxl.load_workbook(r'C:\Users\Salem Magdy\Downloads\الإدارة الفنية (1).xlsx')
ws2 = wb2.active
file2_emps = []
for r in range(2, ws2.max_row + 1):
    name = str(ws2.cell(r, 1).value or '').strip()
    code = str(ws2.cell(r, 2).value or '').strip()
    title = str(ws2.cell(r, 3).value or '').strip()
    loc = str(ws2.cell(r, 4).value or '').strip()
    status_raw = str(ws2.cell(r, 5).value or '').strip()
    status = 'V' if 'جاز' in status_raw else 'P' if 'عمل' in status_raw else status_raw
    if name and code:
        file2_emps.append({'name': name, 'code': code, 'title': title, 'loc': loc, 'status': status})

# Collect unique locations
loc_set = set(e['loc'] for e in file2_emps)
print(f"Total employees in file: {len(file2_emps)}")
print(f"Unique locations: {len(loc_set)}")
print(f"\n=== All locations in the file ===")
for loc in sorted(loc_set):
    count = sum(1 for e in file2_emps if e['loc'] == loc)
    print(f"  '{loc}': {count} employees")
