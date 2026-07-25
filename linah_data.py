#!/usr/bin/env python3
"""
linah_data.py
=============
قاعدة بيانات SQLite للقوة والسكن — تعمل مع LINAHSYSTEM.

الاستخدام:
  python linah_data.py import <ملف_النسخة.json>   # استيراد من تصدير التطبيق
  python linah_data.py export                      # تصدير JSON للاستيراد في التطبيق
  python linah_data.py housing                     # عرض السكن (قطاع → غرف → أسماء الموظفين)
  python linah_data.py rebuild                     # إعادة بناء السكن من بيانات العاملين
  python linah_data.py stats                       # إحصائيات عامة
  python linah_data.py excel                       # تصدير Excel (شيت القوة + شيت السكن)
"""

import json, sys, os, sqlite3
from pathlib import Path
from collections import OrderedDict
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

DB_DIR = Path.home() / "Desktop" / "LINAHSYSTEM"
DB_PATH = DB_DIR / "linah_data.db"


def get_db():
    DB_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS employees (
            id TEXT PRIMARY KEY,
            code TEXT,
            name TEXT,
            contract TEXT DEFAULT 'دائم',
            national_id TEXT,
            hire_date TEXT,
            dept TEXT,
            title TEXT,
            gov TEXT,
            sector TEXT,
            room TEXT,
            status TEXT DEFAULT 'P',
            vacation_balance REAL DEFAULT 30,
            assets TEXT DEFAULT '[]'
        );

        CREATE TABLE IF NOT EXISTS rooms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sector TEXT NOT NULL,
            room_number TEXT NOT NULL,
            beds INTEGER DEFAULT 1,
            UNIQUE(sector, room_number)
        );

        CREATE TABLE IF NOT EXISTS sectors (
            name TEXT PRIMARY KEY
        );

        CREATE TABLE IF NOT EXISTS hospitality (
            id TEXT PRIMARY KEY,
            name TEXT,
            type TEXT,
            title TEXT,
            arrival TEXT,
            departure TEXT,
            guests INTEGER DEFAULT 1,
            meals TEXT DEFAULT '[]'
        );
    """)
    conn.commit()
    conn.close()


def import_from_backup(backup_path):
    with open(backup_path, encoding="utf-8-sig") as f:
        data = json.load(f)

    conn = get_db()
    c = conn.cursor()

    # employees
    emps = data.get("employees") or []
    c.execute("DELETE FROM employees")
    for e in emps:
        c.execute("""
            INSERT OR REPLACE INTO employees
            (id, code, name, contract, national_id, hire_date, dept, title, gov, sector, room, status, vacation_balance, assets)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            e.get("id", ""),
            e.get("code", ""),
            e.get("name", ""),
            e.get("contract", "دائم"),
            e.get("nationalId", ""),
            e.get("hireDate", ""),
            e.get("dept", ""),
            e.get("title", ""),
            e.get("gov", ""),
            e.get("sector", ""),
            e.get("room", ""),
            e.get("status", "P"),
            e.get("vacationBalance", 30),
            json.dumps(e.get("assets", []), ensure_ascii=False)
        ))

    # rooms
    rooms = data.get("roomsCapacity") or []
    c.execute("DELETE FROM rooms")
    for r in rooms:
        try:
            c.execute("INSERT OR IGNORE INTO rooms (sector, room_number, beds) VALUES (?, ?, ?)",
                      (r.get("sector", ""), r.get("number", ""), r.get("beds", 1)))
        except:
            pass

    # sectors
    sectors = data.get("dynamicSectors") or []
    c.execute("DELETE FROM sectors")
    for s in sectors:
        if s and isinstance(s, str):
            try:
                c.execute("INSERT OR IGNORE INTO sectors (name) VALUES (?)", (s,))
            except:
                pass

    # hospitality
    hosp = data.get("hospitalities") or []
    c.execute("DELETE FROM hospitality")
    for h in hosp:
        c.execute("""
            INSERT OR REPLACE INTO hospitality
            (id, name, type, title, arrival, departure, guests, meals)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            h.get("_id", ""),
            h.get("name", ""),
            h.get("type", ""),
            h.get("title", ""),
            h.get("arrival", ""),
            h.get("departure", ""),
            h.get("guests", 1),
            json.dumps(h.get("meals", []), ensure_ascii=False)
        ))

    conn.commit()
    conn.close()

    print(f"✅ استيراد {len(emps)} موظف, {len(rooms)} غرفة, {len(sectors)} قطاع, {len(hosp)} ضيافة")
    rebuild_housing()


def rebuild_housing():
    conn = get_db()
    c = conn.cursor()

    c.execute("DELETE FROM rooms")
    c.execute("DELETE FROM sectors")

    rows = c.execute("SELECT sector, room, COUNT(*) as cnt FROM employees WHERE sector != '' AND room != '' GROUP BY sector, room ORDER BY sector, room").fetchall()
    for r in rows:
        c.execute("INSERT INTO rooms (sector, room_number, beds) VALUES (?, ?, ?)",
                  (r["sector"], r["room"], r["cnt"]))

    sec_rows = c.execute("SELECT DISTINCT sector FROM employees WHERE sector != '' ORDER BY sector").fetchall()
    for r in sec_rows:
        try:
            c.execute("INSERT OR IGNORE INTO sectors (name) VALUES (?)", (r["sector"],))
        except:
            pass

    conn.commit()

    # add missing from rooms table
    sec_from_rooms = c.execute("SELECT DISTINCT sector FROM rooms").fetchall()
    for r in sec_from_rooms:
        try:
            c.execute("INSERT OR IGNORE INTO sectors (name) VALUES (?)", (r["sector"],))
        except:
            pass
    conn.commit()
    conn.close()
    print(f"🏠 تم إعادة بناء السكن: {len(rows)} غرفة/مبنى")


def export_to_json():
    conn = get_db()
    c = conn.cursor()

    employees_out = []
    for r in c.execute("SELECT * FROM employees").fetchall():
        d = dict(r)
        d["nationalId"] = d.pop("national_id", "")
        d["hireDate"] = d.pop("hire_date", "")
        d["vacationBalance"] = d.pop("vacation_balance", 30)
        try:
            d["assets"] = json.loads(d.get("assets", "[]"))
        except:
            d["assets"] = []
        employees_out.append(d)

    rooms_out = [{"sector": r["sector"], "number": r["room_number"], "beds": r["beds"]}
                 for r in c.execute("SELECT * FROM rooms").fetchall()]

    sectors_out = [r["name"] for r in c.execute("SELECT name FROM sectors").fetchall()]

    hosp_out = []
    for r in c.execute("SELECT * FROM hospitality").fetchall():
        d = dict(r)
        d["_id"] = d.pop("id", "")
        d["meals"] = d.pop("meals", "[]")
        try:
            d["meals"] = json.loads(d["meals"])
        except:
            d["meals"] = []
        hosp_out.append(d)

    conn.close()

    output = {
        "employees": employees_out,
        "roomsCapacity": rooms_out,
        "dynamicSectors": sectors_out,
        "hospitalities": hosp_out,
        "_exportedBy": "linah_data.py"
    }

    export_path = DB_DIR / f"linah_export_{__import__('datetime').datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with open(export_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"📦 تم التصدير: {export_path}")
    print(f"   {len(employees_out)} موظف, {len(rooms_out)} غرفة, {len(sectors_out)} قطاع, {len(hosp_out)} ضيافة")


def export_excel():
    if not HAS_OPENPYXL:
        print("❌ openpyxl مش مثبت. شغّل: pip install openpyxl")
        return

    conn = get_db()
    c = conn.cursor()

    wb = openpyxl.Workbook()

    # ========== Sheet 1: القوة ==========
    ws1 = wb.active
    ws1.title = "القوة"

    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    emp_headers = ["الكود", "الاسم رباعي", "نوع التعاقد", "رقم الموبايل", "تاريخ التعيين",
                   "الإدارة", "الوظيفة", "المحافظة", "المبنى السكني", "رقم الغرفة", "الحالة"]
    for ci, h in enumerate(emp_headers, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    emp_data = c.execute("""
        SELECT code, name, contract, national_id, hire_date, dept, title, gov, sector, room, status
        FROM employees ORDER BY name
    """).fetchall()

    for ri, r in enumerate(emp_data, 2):
        vals = [r["code"] or "", r["name"] or "", r["contract"] or "", r["national_id"] or "",
                r["hire_date"] or "", r["dept"] or "", r["title"] or "", r["gov"] or "",
                r["sector"] or "", r["room"] or "",
                "متواجد" if r["status"] == "P" else "في إجازة"]
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ri % 2 == 0:
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 18
    ws1.column_dimensions["G"].width = 18
    ws1.column_dimensions["H"].width = 16
    ws1.column_dimensions["I"].width = 16
    ws1.column_dimensions["J"].width = 14
    ws1.column_dimensions["K"].width = 14

    # ========== Sheet 2: السكن ==========
    ws2 = wb.create_sheet("السكن")

    sec_fill = PatternFill(start_color="FF8F00", end_color="FF8F00", fill_type="solid")
    sec_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)

    room_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    room_font = Font(name="Calibri", bold=True, size=11)

    empty_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    row_num = 1
    sectors = c.execute("SELECT name FROM sectors ORDER BY name").fetchall()
    if not sectors:
        sectors = c.execute("SELECT DISTINCT sector FROM rooms ORDER BY sector").fetchall()

    for sec in sectors:
        sname = sec["name"] if "name" in sec else sec["sector"]

        # title row
        cell = ws2.cell(row=row_num, column=1, value=f"🏢 {sname}")
        cell.font = sec_font
        cell.fill = sec_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        row_num += 1

        rooms = c.execute("SELECT * FROM rooms WHERE sector = ? ORDER BY room_number", (sname,)).fetchall()
        if not rooms:
            cell = ws2.cell(row=row_num, column=1, value="(لا توجد غرف)")
            cell.font = Font(italic=True, color="999999")
            row_num += 1
            continue

        for r in rooms:
            residents = c.execute(
                "SELECT name, code, status FROM employees WHERE sector = ? AND room = ? ORDER BY name",
                (sname, r["room_number"])
            ).fetchall()

            occ = len(residents)
            bed_status = "✅" if occ <= r["beds"] else "⚠️"

            # room header
            cell = ws2.cell(row=row_num, column=1, value=f"🛏️ {r['room_number']}")
            cell.font = room_font
            cell.fill = room_fill
            cell.border = thin_border

            cell = ws2.cell(row=row_num, column=2, value=f"{occ}/{r['beds']} سرير {bed_status}")
            cell.font = room_font
            cell.fill = room_fill
            cell.border = thin_border
            ws2.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=3)
            row_num += 1

            if not residents:
                cell = ws2.cell(row=row_num, column=1, value="  شاغرة")
                cell.font = Font(italic=True, color="999999")
                cell.fill = empty_fill
                ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
                row_num += 1
            else:
                for emp in residents:
                    mark = "📌" if emp["status"] == "P" else "🛫"
                    name_cell = ws2.cell(row=row_num, column=1, value=f"  {mark} {emp['name']}")
                    name_cell.border = thin_border
                    code_cell = ws2.cell(row=row_num, column=4, value=emp["code"] or "")
                    code_cell.border = thin_border
                    st_cell = ws2.cell(row=row_num, column=5,
                                       value="متواجد" if emp["status"] == "P" else "في إجازة")
                    st_cell.border = thin_border
                    if row_num % 2 == 0:
                        for ci in range(1, 6):
                            ws2.cell(row=row_num, column=ci).fill = PatternFill(
                                start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    row_num += 1

        row_num += 1

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14

    conn.close()

    xlsx_path = DB_DIR / f"linah_report_{__import__('datetime').datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(xlsx_path))
    print(f"📗 تم التصدير: {xlsx_path}")
    print(f"   شيت القوة: {len(emp_data)} موظف")
    print(f"   شيت السكن: {len(emp_data)} موظف في {len(rooms)} غرفة")


def show_housing():
    conn = get_db()
    c = conn.cursor()

    sectors = c.execute("SELECT name FROM sectors ORDER BY name").fetchall()
    if not sectors:
        sectors = c.execute("SELECT DISTINCT sector FROM rooms ORDER BY sector").fetchall()

    if not sectors:
        print("⚠️  لا توجد بيانات سكن — استخدم 'import' أولاً أو 'rebuild'")
        conn.close()
        return

    for sec in sectors:
        sname = sec["name"] if "name" in sec else sec["sector"]
        print(f"\n{'='*60}")
        print(f"  🏢 {sname}")
        print(f"{'='*60}")

        rooms = c.execute("SELECT * FROM rooms WHERE sector = ? ORDER BY room_number", (sname,)).fetchall()
        if not rooms:
            print("  (لا توجد غرف)")
            continue

        for r in rooms:
            residents = c.execute(
                "SELECT name, code, status FROM employees WHERE sector = ? AND room = ? ORDER BY name",
                (sname, r["room_number"])
            ).fetchall()

            occ = len(residents)
            status_icon = "✅" if occ <= r["beds"] else "⚠️"
            print(f"\n  🛏️  {r['room_number']}  —  {occ}/{r['beds']} سرير {status_icon}")
            for emp in residents:
                mark = "📌" if emp["status"] == "P" else "🛫"
                print(f"     {mark} {emp['name']}  ({emp['code'] or '—'})")

    conn.close()


def show_stats():
    conn = get_db()
    c = conn.cursor()

    total = c.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
    perm = c.execute("SELECT COUNT(*) FROM employees WHERE contract = 'دائم'").fetchone()[0]
    casual = c.execute("SELECT COUNT(*) FROM employees WHERE contract = 'كاجول'").fetchone()[0]
    present = c.execute("SELECT COUNT(*) FROM employees WHERE status = 'P'").fetchone()[0]
    vacation = c.execute("SELECT COUNT(*) FROM employees WHERE status = 'V'").fetchone()[0]
    housed = c.execute("SELECT COUNT(*) FROM employees WHERE sector != '' AND room != ''").fetchone()[0]
    rooms_count = c.execute("SELECT COUNT(*) FROM rooms").fetchone()[0]
    sectors_count = c.execute("SELECT COUNT(*) FROM sectors").fetchone()[0]

    conn.close()

    print(f"{'='*40}")
    print(f"  📊 إحصائيات قاعدة بيانات LINAHSYSTEM")
    print(f"{'='*40}")
    print(f"  إجمالي القوة:          {total}")
    print(f"  دائم:                  {perm}")
    print(f"  كاجول:                 {casual}")
    print(f"  متواجدون:              {present}")
    print(f"  في إجازة:              {vacation}")
    print(f"  لهم سكن:               {housed}")
    print(f"  عدد الغرف:             {rooms_count}")
    print(f"  عدد القطاعات:          {sectors_count}")
    print(f"{'='*40}")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return

    init_db()
    cmd = sys.argv[1]

    if cmd == "import":
        if len(sys.argv) < 3:
            print("❌ حدد ملف JSON: python linah_data.py import <file.json>")
            return
        import_from_backup(Path(sys.argv[2]))
    elif cmd == "export":
        export_to_json()
    elif cmd == "housing":
        show_housing()
    elif cmd == "rebuild":
        rebuild_housing()
    elif cmd == "stats":
        show_stats()
    elif cmd == "excel":
        export_excel()
    else:
        print(f"❌ أمر غير معروف: {cmd}")
        print(__doc__)


if __name__ == "__main__":
    main()
